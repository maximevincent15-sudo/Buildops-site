"""
Génère un fichier XLSX de test pour l'import clients du SaaS Firovia.

Format respecté :
- Ligne 1 : en-têtes (Nom, Contact, Email, Téléphone, Adresse, Notes internes)
- Ligne 2 : exemple (sera ignorée par le parser via la détection isHintRow)
- Ligne 3+ : 10 clients fictifs crédibles, secteurs variés

Le mapping header est tolérant (accents/casse/espaces ignorés), donc on peut
mettre les libellés tels que définis dans clientsImport.ts.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "clients"

# ─── En-têtes (ligne 1) ──────────────────────────────────────────────
headers = [
    "Nom *",          # obligatoire
    "Contact",
    "Email",
    "Téléphone",
    "Adresse",
    "Notes internes",
]

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1C2130")
    cell.alignment = Alignment(vertical="center", horizontal="left")
    cell.border = Border(bottom=Side(style="thin", color="000000"))

ws.row_dimensions[1].height = 22

# ─── Ligne 2 : exemple (sera ignorée par le parser) ────────────────
example_row = [
    "ex: Valoris SA",
    "ex: M. Dupont",
    "ex: contact@valoris.fr",
    "ex: 0123456789",
    "ex: 12 rue des Lilas, 75010 Paris",
    "ex: parking sous-sol",
]
for col_idx, val in enumerate(example_row, start=1):
    cell = ws.cell(row=2, column=col_idx, value=val)
    cell.font = Font(name="Arial", italic=True, color="9AA0AE", size=10)

# ─── Ligne 3+ : 10 clients fictifs crédibles ──────────────────────
# Secteurs variés (ERP, IGH, résidentiel collectif, tertiaire, industriel)
# Tous noms/contacts/emails/adresses entièrement fictifs.
clients = [
    [
        "Carrelages Méditerranée",
        "Mme Lefranc",
        "contact@carrelages-mediterranee.fr",
        "04 91 56 78 90",
        "27 Avenue Robert Schuman, 13002 Marseille",
        "Showroom + atelier, 2 niveaux. Demander Mme Lefranc.",
    ],
    [
        "Bâti-Sud Construction",
        "M. Lemoine",
        "j.lemoine@bati-sud.fr",
        "05 61 42 87 33",
        "8 Boulevard de l'Embouchure, 31200 Toulouse",
        "Siège social. Plusieurs chantiers en région.",
    ],
    [
        "Résidence Les Promenades de l'Atlantique",
        "Syndic Atlantique",
        "syndic@promenades-atlantique.fr",
        "02 40 12 34 56",
        "15 Boulevard de la Côte d'Argent, 44600 Saint-Nazaire",
        "Résidence 4 bâtiments, 80 logements. Parking souterrain.",
    ],
    [
        "Hôtel des Voyageurs",
        "M. Bertrand",
        "direction@hotel-voyageurs-strasbourg.fr",
        "03 88 32 19 45",
        "12 Place de la Gare, 67000 Strasbourg",
        "Hôtel 3*, 42 chambres. ERP type O.",
    ],
    [
        "EHPAD Les Lavandes",
        "Mme Hernandez",
        "direction@ehpad-lavandes.fr",
        "04 75 23 45 67",
        "3 Chemin des Cyprès, 26000 Valence",
        "85 résidents. Sensible — prévenir avant intervention.",
    ],
    [
        "Lycée Professionnel Léonard de Vinci",
        "Service maintenance",
        "maintenance@lyc-vinci-rennes.fr",
        "02 23 45 67 89",
        "44 Rue de la Borderie, 35000 Rennes",
        "ERP type R. Intervention hors vacances scolaires.",
    ],
    [
        "Imprimerie Régionale du Nord",
        "M. Vandenberghe",
        "p.vandenberghe@imprimerie-nord.fr",
        "03 20 67 89 12",
        "Zone d'Activité La Pilaterie, 59290 Wasquehal",
        "Stockage papier important. ERP type S. Risque feu élevé.",
    ],
    [
        "Garage Auto-Méca Toulouse",
        "M. Caraballo",
        "contact@automeca-toulouse.fr",
        "05 61 78 90 12",
        "112 Route de Narbonne, 31400 Toulouse",
        "Atelier mécanique + carrosserie. Stockage huiles/solvants.",
    ],
    [
        "Boulangerie Industrielle Pichon",
        "Mme Pichon",
        "compta@boulangerie-pichon.fr",
        "04 78 34 56 78",
        "Avenue des Industries, 69800 Saint-Priest",
        "Production 24/7. Intervention de nuit (créneau 1h-5h).",
    ],
    [
        "Centre Commercial Les Portes du Soleil",
        "Direction technique",
        "technique@portes-du-soleil.fr",
        "04 50 23 45 67",
        "Route d'Annecy, 74330 Épagny",
        "ERP type M. 32 boutiques + parking 600 places.",
    ],
]

for row_idx, row_data in enumerate(clients, start=3):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row_idx].height = 30

# ─── Largeurs de colonnes ──────────────────────────────────────────
column_widths = [38, 22, 38, 18, 50, 50]
for col_idx, width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ─── Freeze panes (header figé) ───────────────────────────────────
ws.freeze_panes = "A2"

# ─── Sauvegarde ────────────────────────────────────────────────────
output_path = "/Users/macbookair/Saas BTP/scripts/clients-test-firovia.xlsx"
wb.save(output_path)
print(f"✅ Fichier généré : {output_path}")
print(f"   {len(clients)} clients prêts à importer.")
